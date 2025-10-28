from __future__ import annotations

from typing import Any, Optional

from openpyxl.worksheet.worksheet import Worksheet


def get_column_index_by_name(
    worksheet: Worksheet,
    column_name: str,
    *,
    header_row: int = 1,
    case_sensitive: bool = True,
) -> Optional[int]:
    """Return the 1-based column index for the given header name.

    Args:
        worksheet: OpenPyXL worksheet to read.
        column_name: Header text to match.
        header_row: 1-based row index containing headers. Defaults to 1.
        case_sensitive: When False, header comparison ignores case.

    Returns:
        The 1-based column index if found, otherwise ``None``.
    """
    if header_row < 1:
        raise ValueError("header_row 必须大于等于 1")

    target = column_name if case_sensitive else column_name.lower()

    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row, column=col).value
        if cell_value is None:
            continue

        candidate = cell_value if case_sensitive else str(cell_value).lower()
        if candidate == target:
            return col

    return None


def get_cell_value(
    worksheet: Worksheet,
    row_index: int,
    column_name: str,
    *,
    header_row: int = 1,
    case_sensitive: bool = True,
) -> Any:
    """Get a cell value by row index and header name."""
    if row_index < 1:
        raise ValueError("row_index 必须大于等于 1")

    col_idx = get_column_index_by_name(
        worksheet,
        column_name,
        header_row=header_row,
        case_sensitive=case_sensitive,
    )
    if col_idx is None:
        return None
    return worksheet.cell(row=row_index, column=col_idx).value


def set_cell_value(
    worksheet: Worksheet,
    row_index: int,
    column_name: str,
    value: Any,
    *,
    header_row: int = 1,
    case_sensitive: bool = True,
) -> bool:
    """Set a cell value by row index and header name."""
    if row_index < 1:
        raise ValueError("row_index 必须大于等于 1")

    col_idx = get_column_index_by_name(
        worksheet,
        column_name,
        header_row=header_row,
        case_sensitive=case_sensitive,
    )
    if col_idx is None:
        return False

    worksheet.cell(row=row_index, column=col_idx).value = value
    return True


__all__ = [
    "get_column_index_by_name",
    "get_cell_value",
    "set_cell_value",
]
